import pandas as pd
import numpy as np
import openpyxl

folder = '/Users/gengyunxin/Documents/项目/traffic_model/TVTS_9.28/finetune/test_result_all/prediction_plot/'
week_len = 288*7
# for word_len in [1,2,6,12,24,'lstm']:
for word_len in ['transformer']:
    for prediction_len in [1,6,12,36,72]:
        if prediction_len==1:
            start = 71
        elif prediction_len==6:
            start = 66
        elif prediction_len==12:
            start = 60
        elif prediction_len==36:
            start = 36
        elif prediction_len==72:
            start = 0

        file = folder + f'{word_len}_{prediction_len}_plot.csv'
        save_path1 = folder + f'predict_aweek_excel/{word_len}_{prediction_len}.xlsx'
        save_path2 = folder + f'predict_aweek_excel/predict_2016/{word_len}_{prediction_len}.xlsx'
        df = pd.read_csv(file)
        df_head = df.iloc[start:]
        df_tail = df_head.iloc[:week_len] # 起始按P72对齐17.126684，长度2016
        df_tail.to_excel(save_path1, index=False, columns=['Input', 'Predict'])

        wb = openpyxl.load_workbook(save_path1)
        ws = wb.active
        for row in range(2, 290):
            ws.cell(row=row, column=3, value=ws.cell(row=row+1152, column=1).value)
            ws.cell(row=row, column=4, value=ws.cell(row=row+1152, column=2).value)
        wb.save(save_path2)

